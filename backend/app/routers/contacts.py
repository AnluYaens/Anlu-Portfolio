from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from io import BytesIO
from sqlmodel import Session, select
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..database import get_session
from ..models import ContactMessage
from ..schemas import ContactCreate
from ..security import verify_admin
from ..limiter import limiter

router = APIRouter(prefix="/contacts", tags=["contacts"])

@router.post("/", response_model=ContactMessage)
@limiter.limit("5/minute")
def create_contact(request: Request, contact: ContactCreate, session: Session = Depends(get_session)):
    db_contact = ContactMessage.model_validate(contact)
    session.add(db_contact)
    session.commit()
    session.refresh(db_contact)
    return db_contact

@router.get("/", response_model=List[ContactMessage], dependencies=[Depends(verify_admin)])
@limiter.limit("30/minute")
def read_contacts(request: Request, session: Session = Depends(get_session)):
    contacts = session.exec(select(ContactMessage).order_by(ContactMessage.created_at.desc())).all()
    return contacts

@router.get("/export-excel", dependencies=[Depends(verify_admin)])
@limiter.limit("5/minute")
def export_contacts_excel(request: Request, session: Session = Depends(get_session)):
    contacts = session.exec(select(ContactMessage).order_by(ContactMessage.created_at.desc())).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Contact Messages"
    
    # Headers
    headers = ["ID", "Name", "Email", "Message", "Date"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for contact in contacts:
        created_at = contact.created_at.strftime("%Y-%m-%d %H:%M:%S") if contact.created_at else ""
        ws.append([contact.id, contact.name, contact.email, contact.message, created_at])
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50) # Max width 50
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {"Content-Disposition": "attachment; filename=contacts.xlsx"}
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
